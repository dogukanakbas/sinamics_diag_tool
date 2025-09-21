"""
Advanced Reporting System - Gelişmiş raporlama sistemi
"""
import os
import json
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


class PDFReportGenerator:
    """PDF rapor oluşturucu"""
    
    def __init__(self):
        self.template_path = "reporting/templates/"
        self.output_path = "reports/"
        self._ensure_directories()
        
    def _ensure_directories(self):
        """Gerekli klasörleri oluştur"""
        os.makedirs(self.template_path, exist_ok=True)
        os.makedirs(self.output_path, exist_ok=True)
        
    def generate_diagnostic_report(self, data: Dict[str, Any], 
                                 report_type: str = "daily") -> str:
        """Teşhis raporu oluştur"""
        try:
            # HTML template oluştur
            html_content = self._create_html_template(data, report_type)
            
            # PDF'e çevir (weasyprint kullanarak)
            pdf_filename = f"diagnostic_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            pdf_path = os.path.join(self.output_path, pdf_filename)
            
            # Basit HTML to PDF conversion (gerçek uygulamada weasyprint kullanılmalı)
            self._html_to_pdf(html_content, pdf_path)
            
            return pdf_path
            
        except Exception as e:
            print(f"PDF generation error: {e}")
            return None
            
    def _create_html_template(self, data: Dict[str, Any], report_type: str) -> str:
        """HTML template oluştur"""
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>SINAMICS Diagnostic Report</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                .header {{ text-align: center; border-bottom: 2px solid #333; padding-bottom: 20px; }}
                .section {{ margin: 20px 0; }}
                .section h2 {{ color: #333; border-bottom: 1px solid #ccc; }}
                table {{ width: 100%; border-collapse: collapse; margin: 10px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
                .fault {{ color: #d32f2f; }}
                .alarm {{ color: #f57c00; }}
                .normal {{ color: #388e3c; }}
                .summary {{ background-color: #f5f5f5; padding: 15px; border-radius: 5px; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>SINAMICS Diagnostic Report</h1>
                <p>Report Type: {report_type.title()}</p>
                <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
            
            <div class="section">
                <h2>Executive Summary</h2>
                <div class="summary">
                    <p><strong>Total Events:</strong> {data.get('summary', {}).get('total_events', 0)}</p>
                    <p><strong>Total Faults:</strong> {data.get('summary', {}).get('total_faults', 0)}</p>
                    <p><strong>Total Alarms:</strong> {data.get('summary', {}).get('total_alarms', 0)}</p>
                    <p><strong>Daily Average:</strong> {data.get('summary', {}).get('daily_average', 0)}</p>
                </div>
            </div>
            
            <div class="section">
                <h2>Fault Analysis</h2>
                <table>
                    <tr>
                        <th>Fault ID</th>
                        <th>Component</th>
                        <th>Description</th>
                        <th>Severity</th>
                        <th>Timestamp</th>
                    </tr>
        """
        
        # Fault'ları ekle
        for fault in data.get('faults', []):
            html += f"""
                    <tr class="fault">
                        <td>{fault.get('id', 'N/A')}</td>
                        <td>{fault.get('component', 'N/A')}</td>
                        <td>{fault.get('desc', 'N/A')}</td>
                        <td>{fault.get('severity', 'N/A')}</td>
                        <td>{fault.get('timestamp', 'N/A')}</td>
                    </tr>
            """
            
        html += """
                </table>
            </div>
            
            <div class="section">
                <h2>Alarm Analysis</h2>
                <table>
                    <tr>
                        <th>Alarm ID</th>
                        <th>Component</th>
                        <th>Description</th>
                        <th>Severity</th>
                        <th>Status</th>
                        <th>Timestamp</th>
                    </tr>
        """
        
        # Alarm'ları ekle
        for alarm in data.get('alarms', []):
            html += f"""
                    <tr class="alarm">
                        <td>{alarm.get('id', 'N/A')}</td>
                        <td>{alarm.get('component', 'N/A')}</td>
                        <td>{alarm.get('desc', 'N/A')}</td>
                        <td>{alarm.get('severity', 'N/A')}</td>
                        <td>{alarm.get('status', 'N/A')}</td>
                        <td>{alarm.get('timestamp', 'N/A')}</td>
                    </tr>
            """
            
        html += """
                </table>
            </div>
            
            <div class="section">
                <h2>Component Analysis</h2>
                <table>
                    <tr>
                        <th>Component</th>
                        <th>Faults</th>
                        <th>Alarms</th>
                        <th>Total Issues</th>
                    </tr>
        """
        
        # Bileşen analizi
        component_stats = data.get('component_analysis', {})
        for component, stats in component_stats.items():
            total_issues = stats.get('faults', 0) + stats.get('alarms', 0)
            html += f"""
                    <tr>
                        <td>{component}</td>
                        <td>{stats.get('faults', 0)}</td>
                        <td>{stats.get('alarms', 0)}</td>
                        <td>{total_issues}</td>
                    </tr>
            """
            
        html += """
                </table>
            </div>
            
            <div class="section">
                <h2>Recommendations</h2>
                <ul>
        """
        
        # Öneriler
        recommendations = data.get('recommendations', [])
        for rec in recommendations:
            html += f"<li>{rec}</li>"
            
        html += """
                </ul>
            </div>
            
            <div class="section">
                <p><em>Report generated by SINAMICS Diag & Viz System</em></p>
            </div>
        </body>
        </html>
        """
        
        return html
        
    def _html_to_pdf(self, html_content: str, output_path: str):
        """HTML'i PDF'e çevir"""
        # Basit HTML to PDF conversion
        # Gerçek uygulamada weasyprint veya wkhtmltopdf kullanılmalı
        try:
            # HTML dosyası olarak kaydet
            html_path = output_path.replace('.pdf', '.html')
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
                
            # PDF'e çevir (basit yöntem)
            import subprocess
            try:
                # wkhtmltopdf kullanarak PDF oluştur
                subprocess.run(['wkhtmltopdf', html_path, output_path], 
                             check=True, capture_output=True)
            except (subprocess.CalledProcessError, FileNotFoundError):
                # wkhtmltopdf yoksa HTML dosyasını PDF olarak kopyala
                with open(output_path, 'w', encoding='utf-8') as f:
                    f.write(html_content)
                    
        except Exception as e:
            print(f"HTML to PDF conversion error: {e}")


class ExcelReportGenerator:
    """Excel rapor oluşturucu"""
    
    def __init__(self):
        self.output_path = "reports/"
        os.makedirs(self.output_path, exist_ok=True)
        
    def generate_diagnostic_report(self, data: Dict[str, Any], 
                                 report_type: str = "daily") -> str:
        """Excel teşhis raporu oluştur"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Yeni workbook oluştur
            wb = openpyxl.Workbook()
            
            # Summary sheet
            self._create_summary_sheet(wb, data)
            
            # Faults sheet
            self._create_faults_sheet(wb, data)
            
            # Alarms sheet
            self._create_alarms_sheet(wb, data)
            
            # Component analysis sheet
            self._create_component_analysis_sheet(wb, data)
            
            # Dosyayı kaydet
            excel_filename = f"diagnostic_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            excel_path = os.path.join(self.output_path, excel_filename)
            wb.save(excel_path)
            
            return excel_path
            
        except ImportError:
            print("openpyxl not installed. Install with: pip install openpyxl")
            return None
        except Exception as e:
            print(f"Excel generation error: {e}")
            return None
            
    def _create_summary_sheet(self, wb, data: Dict[str, Any]):
        """Özet sheet'i oluştur"""
        ws = wb.active
        ws.title = "Summary"
        
        # Başlık
        ws['A1'] = "SINAMICS Diagnostic Report Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Tarih
        ws['A3'] = "Generated:"
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Özet veriler
        summary = data.get('summary', {})
        row = 5
        
        summary_data = [
            ("Total Events", summary.get('total_events', 0)),
            ("Total Faults", summary.get('total_faults', 0)),
            ("Total Alarms", summary.get('total_alarms', 0)),
            ("Daily Average", summary.get('daily_average', 0)),
            ("Last 24h Events", summary.get('last_24h_events', 0))
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
            
        # En problemli bileşen
        most_problematic = data.get('most_problematic_component')
        if most_problematic:
            ws[f'A{row}'] = "Most Problematic Component:"
            ws[f'B{row}'] = most_problematic
            row += 1
            
        # Öneriler
        recommendations = data.get('recommendations', [])
        if recommendations:
            ws[f'A{row}'] = "Recommendations:"
            row += 1
            for rec in recommendations:
                ws[f'A{row}'] = f"• {rec}"
                row += 1
                
    def _create_faults_sheet(self, wb, data: Dict[str, Any]):
        """Fault'lar sheet'i oluştur"""
        ws = wb.create_sheet("Faults")
        
        # Başlıklar
        headers = ["Fault ID", "Component", "Description", "Severity", "Timestamp"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
        # Fault verileri
        faults = data.get('faults', [])
        for row, fault in enumerate(faults, 2):
            ws.cell(row=row, column=1, value=fault.get('id', ''))
            ws.cell(row=row, column=2, value=fault.get('component', ''))
            ws.cell(row=row, column=3, value=fault.get('desc', ''))
            ws.cell(row=row, column=4, value=fault.get('severity', ''))
            ws.cell(row=row, column=5, value=fault.get('timestamp', ''))
            
    def _create_alarms_sheet(self, wb, data: Dict[str, Any]):
        """Alarm'lar sheet'i oluştur"""
        ws = wb.create_sheet("Alarms")
        
        # Başlıklar
        headers = ["Alarm ID", "Component", "Description", "Severity", "Status", "Timestamp"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
        # Alarm verileri
        alarms = data.get('alarms', [])
        for row, alarm in enumerate(alarms, 2):
            ws.cell(row=row, column=1, value=alarm.get('id', ''))
            ws.cell(row=row, column=2, value=alarm.get('component', ''))
            ws.cell(row=row, column=3, value=alarm.get('desc', ''))
            ws.cell(row=row, column=4, value=alarm.get('severity', ''))
            ws.cell(row=row, column=5, value=alarm.get('status', ''))
            ws.cell(row=row, column=6, value=alarm.get('timestamp', ''))
            
    def _create_component_analysis_sheet(self, wb, data: Dict[str, Any]):
        """Bileşen analizi sheet'i oluştur"""
        ws = wb.create_sheet("Component Analysis")
        
        # Başlıklar
        headers = ["Component", "Faults", "Alarms", "Total Issues", "Percentage"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
        # Bileşen verileri
        component_stats = data.get('component_analysis', {})
        total_issues = sum(stats.get('faults', 0) + stats.get('alarms', 0) 
                          for stats in component_stats.values())
        
        for row, (component, stats) in enumerate(component_stats.items(), 2):
            faults = stats.get('faults', 0)
            alarms = stats.get('alarms', 0)
            total = faults + alarms
            percentage = (total / total_issues * 100) if total_issues > 0 else 0
            
            ws.cell(row=row, column=1, value=component)
            ws.cell(row=row, column=2, value=faults)
            ws.cell(row=row, column=3, value=alarms)
            ws.cell(row=row, column=4, value=total)
            ws.cell(row=row, column=5, value=f"{percentage:.1f}%")


class EmailReportSender:
    """E-posta rapor gönderici"""
    
    def __init__(self, smtp_server: str = "smtp.gmail.com", smtp_port: int = 587):
        self.smtp_server = smtp_server
        self.smtp_port = smtp_port
        self.username = None
        self.password = None
        
    def configure(self, username: str, password: str):
        """E-posta konfigürasyonu"""
        self.username = username
        self.password = password
        
    def send_report(self, recipients: List[str], subject: str, 
                   body: str, attachments: List[str] = None) -> bool:
        """Rapor gönder"""
        if not self.username or not self.password:
            print("Email not configured")
            return False
            
        try:
            # E-posta mesajı oluştur
            msg = MIMEMultipart()
            msg['From'] = self.username
            msg['To'] = ', '.join(recipients)
            msg['Subject'] = subject
            
            # Mesaj gövdesi
            msg.attach(MIMEText(body, 'html'))
            
            # Ekler
            if attachments:
                for attachment_path in attachments:
                    if os.path.exists(attachment_path):
                        with open(attachment_path, "rb") as attachment:
                            part = MIMEBase('application', 'octet-stream')
                            part.set_payload(attachment.read())
                            
                        encoders.encode_base64(part)
                        part.add_header(
                            'Content-Disposition',
                            f'attachment; filename= {os.path.basename(attachment_path)}'
                        )
                        msg.attach(part)
                        
            # E-posta gönder
            server = smtplib.SMTP(self.smtp_server, self.smtp_port)
            server.starttls()
            server.login(self.username, self.password)
            text = msg.as_string()
            server.sendmail(self.username, recipients, text)
            server.quit()
            
            return True
            
        except Exception as e:
            print(f"Email sending error: {e}")
            return False


class AdvancedReportingManager:
    """Gelişmiş raporlama yöneticisi"""
    
    def __init__(self):
        self.pdf_generator = PDFReportGenerator()
        self.excel_generator = ExcelReportGenerator()
        self.email_sender = EmailReportSender()
        
    def generate_comprehensive_report(self, history_data: List[Dict[str, Any]], 
                                    report_type: str = "daily") -> Dict[str, Any]:
        """Kapsamlı rapor oluştur"""
        # Rapor verilerini hazırla
        report_data = self._prepare_report_data(history_data, report_type)
        
        # PDF raporu oluştur
        pdf_path = self.pdf_generator.generate_diagnostic_report(report_data, report_type)
        
        # Excel raporu oluştur
        excel_path = self.excel_generator.generate_diagnostic_report(report_data, report_type)
        
        return {
            'report_data': report_data,
            'pdf_path': pdf_path,
            'excel_path': excel_path,
            'generated_at': datetime.now().isoformat()
        }
        
    def _prepare_report_data(self, history_data: List[Dict[str, Any]], 
                           report_type: str) -> Dict[str, Any]:
        """Rapor verilerini hazırla"""
        # Tarih aralığını belirle
        if report_type == "daily":
            start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = start_date + timedelta(days=1)
        elif report_type == "weekly":
            start_date = datetime.now() - timedelta(days=7)
            end_date = datetime.now()
        else:  # monthly
            start_date = datetime.now() - timedelta(days=30)
            end_date = datetime.now()
            
        # Verileri filtrele
        filtered_data = [
            entry for entry in history_data
            if start_date <= datetime.fromisoformat(entry['timestamp']) <= end_date
        ]
        
        # Temel istatistikler
        total_events = len(filtered_data)
        faults = [e for e in filtered_data if e['type'] == 'fault']
        alarms = [e for e in filtered_data if e['type'] == 'alarm']
        
        # Bileşen analizi
        component_stats = {}
        for entry in filtered_data:
            if entry['type'] in ['fault', 'alarm']:
                component = entry['data'].get('component', 'Unknown')
                if component not in component_stats:
                    component_stats[component] = {'faults': 0, 'alarms': 0}
                component_stats[component][entry['type'] + 's'] += 1
                
        # En problemli bileşen
        most_problematic = max(component_stats.items(), 
                             key=lambda x: x[1]['faults'] + x[1]['alarms']) if component_stats else None
        
        # Günlük ortalama
        days = (end_date - start_date).days or 1
        daily_avg = total_events / days
        
        # Son 24 saat
        last_24h = datetime.now() - timedelta(hours=24)
        recent_events = len([e for e in filtered_data 
                           if datetime.fromisoformat(e['timestamp']) > last_24h])
        
        # Öneriler
        recommendations = self._generate_recommendations(component_stats, most_problematic)
        
        return {
            'summary': {
                'total_events': total_events,
                'total_faults': len(faults),
                'total_alarms': len(alarms),
                'daily_average': round(daily_avg, 2),
                'last_24h_events': recent_events
            },
            'faults': [f['data'] for f in faults],
            'alarms': [a['data'] for a in alarms],
            'component_analysis': component_stats,
            'most_problematic_component': most_problematic[0] if most_problematic else None,
            'recommendations': recommendations
        }
        
    def _generate_recommendations(self, component_stats: Dict, 
                                most_problematic: Optional[tuple]) -> List[str]:
        """Öneriler oluştur"""
        recommendations = []
        
        if most_problematic:
            component, stats = most_problematic
            total_issues = stats['faults'] + stats['alarms']
            
            if total_issues > 10:
                recommendations.append(f"High priority: {component} component has {total_issues} issues")
            elif total_issues > 5:
                recommendations.append(f"Medium priority: {component} component needs attention")
                
        # Genel öneriler
        if len(component_stats) > 0:
            avg_issues = sum(s['faults'] + s['alarms'] for s in component_stats.values()) / len(component_stats)
            if avg_issues > 3:
                recommendations.append("Consider preventive maintenance schedule")
                
        if not recommendations:
            recommendations.append("System is operating normally")
            
        return recommendations
        
    def show_reporting_dialog(self, parent, history_data: List[Dict[str, Any]]):
        """Raporlama dialog'u göster"""
        dialog = tk.Toplevel(parent)
        dialog.title("Advanced Reporting")
        dialog.geometry("600x500")
        dialog.transient(parent)
        dialog.grab_set()
        
        # Ana frame
        main_frame = ttk.Frame(dialog, padding="10")
        main_frame.pack(fill="both", expand=True)
        
        # Başlık
        title_label = ttk.Label(main_frame, text="Advanced Diagnostic Reports", 
                               font=("Arial", 14, "bold"))
        title_label.pack(pady=(0, 20))
        
        # Rapor seçenekleri
        options_frame = ttk.LabelFrame(main_frame, text="Report Options", padding="10")
        options_frame.pack(fill="x", pady=(0, 20))
        
        # Rapor tipi
        ttk.Label(options_frame, text="Report Type:").pack(anchor="w")
        report_type_var = tk.StringVar(value="daily")
        report_type_combo = ttk.Combobox(options_frame, textvariable=report_type_var,
                                        values=["daily", "weekly", "monthly"],
                                        state="readonly", width=20)
        report_type_combo.pack(anchor="w", pady=(5, 15))
        
        # Format seçenekleri
        format_frame = ttk.Frame(options_frame)
        format_frame.pack(fill="x")
        
        pdf_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(format_frame, text="PDF Report", variable=pdf_var).pack(anchor="w")
        
        excel_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(format_frame, text="Excel Report", variable=excel_var).pack(anchor="w")
        
        email_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(format_frame, text="Send via Email", variable=email_var).pack(anchor="w")
        
        # E-posta ayarları
        email_frame = ttk.LabelFrame(main_frame, text="Email Settings", padding="10")
        email_frame.pack(fill="x", pady=(0, 20))
        
        ttk.Label(email_frame, text="Recipients (comma separated):").pack(anchor="w")
        recipients_var = tk.StringVar()
        recipients_entry = ttk.Entry(email_frame, textvariable=recipients_var, width=50)
        recipients_entry.pack(fill="x", pady=(5, 10))
        
        # Butonlar
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=(20, 0))
        
        def generate_reports():
            try:
                report_type = report_type_var.get()
                
                # Rapor oluştur
                result = self.generate_comprehensive_report(history_data, report_type)
                
                generated_files = []
                if pdf_var.get() and result['pdf_path']:
                    generated_files.append(result['pdf_path'])
                if excel_var.get() and result['excel_path']:
                    generated_files.append(result['excel_path'])
                    
                if generated_files:
                    messagebox.showinfo("Success", 
                                      f"Generated {len(generated_files)} report(s)")
                    
                    # E-posta gönder
                    if email_var.get() and recipients_var.get():
                        recipients = [r.strip() for r in recipients_var.get().split(',')]
                        subject = f"SINAMICS Diagnostic Report - {report_type.title()}"
                        body = f"Please find attached the {report_type} diagnostic report."
                        
                        if self.email_sender.send_report(recipients, subject, body, generated_files):
                            messagebox.showinfo("Success", "Reports sent via email")
                        else:
                            messagebox.showerror("Error", "Failed to send email")
                else:
                    messagebox.showwarning("Warning", "No reports were generated")
                    
            except Exception as e:
                messagebox.showerror("Error", f"Failed to generate reports: {e}")
            finally:
                dialog.destroy()
                
        ttk.Button(button_frame, text="Generate Reports", 
                  command=generate_reports).pack(side="left", padx=(0, 10))
        ttk.Button(button_frame, text="Cancel", 
                  command=dialog.destroy).pack(side="left")
        
        # Status
        status_label = ttk.Label(main_frame, text=f"Total history entries: {len(history_data)}")
        status_label.pack(pady=(20, 0))


# Global reporting manager
advanced_reporting = AdvancedReportingManager()
